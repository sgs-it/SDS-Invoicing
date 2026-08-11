"""Phase 1 — convert the Excel invoicing procedures into structured master data.

Reads `SDS Invoicing_Procedure_.xlsx` and upserts Clients, Projects,
ProjectDocumentRequirement rows (required document checklist) and workflow
metadata (criteria, step narrative, responsible team) into the SQLite database.

Usage:
    python migrate_import.py            # import (idempotent, skips existing)
    python migrate_import.py --reset    # wipe the database first

Run after `pip install -r requirements.txt`.
"""
import argparse
import os
import re
import sys

import openpyxl

from web import create_app
from web.db import db
from web.models import (
    Client,
    DocumentType,
    Project,
    ProjectDocumentRequirement,
    SubmissionMethod,
)

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "SDS Invoicing_Procedure_.xlsx")

# doc code -> (keyword, regex-ish pattern). Ordered so specific phrases win.
DOC_CHECKS = [
    ("KPI", re.compile(r"\bKPI\b", re.I)),
    ("WDA", re.compile(r"\bWDA\b|work done accept", re.I)),
    ("IAF", re.compile(r"\bIAF\b|invoice approval form", re.I)),
    ("GRN", re.compile(r"\bGRN\b|goods receipt", re.I)),
    ("SDS", re.compile(r"\bSDS\b|service delivery sheet", re.I)),
    ("PO", re.compile(r"\bPO\b|purchase order", re.I)),
    ("MR", re.compile(r"monthly report", re.I)),
    ("ATT", re.compile(r"attendance", re.I)),
    ("SR", re.compile(r"service report", re.I)),
    ("DO", re.compile(r"\bDO\b|delivery order", re.I)),
    ("SSHEET", re.compile(r"service sheet", re.I)),
    ("CF", re.compile(r"comp\w*ion", re.I)),  # catches completion/compleltion/complteion
    ("INVOICE", re.compile(r"invoice", re.I)),
]
INVOICE_NOT_REQUIRED = re.compile(r"invoice\s+copy\s+not\s+required", re.I)

TEAM_PATTERNS = [
    ("CSD", re.compile(r"CSD|IP DXB", re.I)),
    ("SGS", re.compile(r"SGS", re.I)),
    ("BPO", re.compile(r"IP BPO|\bBPO\b", re.I)),
    ("OPS", re.compile(r"Operations|\bDXB\b|KPI request", re.I)),
    ("FIN", re.compile(r"Finance", re.I)),
    ("MGMT", re.compile(r"Management", re.I)),
]


def clean(value):
    """Normalise a cell: collapse whitespace / NBSP, drop None."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def detect_docs(text):
    """Return ordered document-type codes present in the required-docs cell."""
    if not text:
        return []
    hits = []
    for code, pattern in DOC_CHECKS:
        m = pattern.search(text)
        if m:
            if code == "INVOICE" and INVOICE_NOT_REQUIRED.search(text):
                continue  # e.g. MAF: "Invoice copy not required"
            hits.append((m.start(), code))
    hits.sort(key=lambda t: t[0])
    return [code for _, code in hits]


def detect_team(text):
    """Return the dept code for the first responsible team found in text."""
    if not text:
        return None
    best = None
    best_pos = None
    for code, pattern in TEAM_PATTERNS:
        m = pattern.search(text)
        if m and (best_pos is None or m.start() < best_pos):
            best, best_pos = code, m.start()
    return best


def extract_client(name):
    """Group site names by their client (text before the first ' - ')."""
    name = clean(name)
    if " - " in name:
        return name.split(" - ")[0].strip()
    return name


def build_narrative(pairs):
    """pairs: list of (step_text, responsible_text). Returns readable narrative."""
    parts = []
    for i, (step, resp) in enumerate(pairs, start=1):
        if not step and not resp:
            continue
        piece = f"Step {i}: {step}" if step else f"Step {i}: (no text)"
        if resp:
            piece += f" — {resp}"
        parts.append(piece)
    return "\n".join(parts)


# --------------------------------------------------------------------------- #
# Per-sheet extractors -> yield dicts describing a project row
# --------------------------------------------------------------------------- #
def rows_direct(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        _, client, job, docs, common, sds, sub, step1, resp1 = r[:9]
        yield {
            "client": client, "job": job, "name": client,
            "docs_text": docs, "common": common, "sds": sds,
            "method": "DIRECT", "target": "",
            "narrative": build_narrative([(step1, resp1)]),
        }


def rows_portal(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        (_, client, job, docs, common, sds, step1, resp1,
         step2, resp2, sub) = r[:11]
        yield {
            "client": client, "job": job, "name": client,
            "docs_text": docs, "common": common, "sds": sds,
            "method": "PORTAL", "target": "",
            "narrative": build_narrative([(step1, resp1), (step2, resp2)]),
        }


def rows_email(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        (_, client, job, docs, common, sds, sub, step1, resp1,
         step2, resp2) = r[:11]
        yield {
            "client": client, "job": job, "name": client,
            "docs_text": docs, "common": common, "sds": sds,
            "method": "EMAIL", "target": "",
            "narrative": build_narrative([(step1, resp1), (step2, resp2)]),
        }


def rows_email_direct(ws):
    # Rows 3+ omit the docs column; inherit docs + narrative from the first row.
    inherited_docs, inherited_narrative = "", ""
    for r in ws.iter_rows(min_row=2, values_only=True):
        (_, client, job, docs, sds, step1, resp1, step2, resp2,
         step3, resp3) = r[:11]
        if docs:
            inherited_docs = docs
        narrative = build_narrative(
            [(step1, resp1), (step2, resp2), (step3, resp3)]
        )
        if narrative:
            inherited_narrative = narrative
        else:
            narrative = inherited_narrative
        yield {
            "client": client, "job": job, "name": client,
            "docs_text": inherited_docs, "common": "", "sds": sds,
            "method": "EMAIL_DIRECT", "target": "",
            "narrative": narrative,
        }


def rows_common_grn(ws):
    for r in ws.iter_rows(min_row=2, values_only=True):
        _, client, docs, target = r[:4]
        rest = list(r[4:])  # step/responsible alternating cells E..Q
        pairs = []
        for i in range(0, len(rest) - 1, 2):
            step, resp = rest[i], rest[i + 1]
            if step or resp:
                pairs.append((step, resp))
        yield {
            "client": client, "job": "", "name": client,
            "docs_text": docs, "common": "GRN Required", "sds": "",
            "method": "COMMON_GRN", "target": target,
            "narrative": build_narrative(pairs),
        }


SHEET_EXTRACTORS = [
    ("Direct submission ", rows_direct),
    ("Online portal ", rows_portal),
    ("Online email ", rows_email),
    ("Online email - Direct", rows_email_direct),
    ("COMMON GRN  Invoicing", rows_common_grn),
]


# --------------------------------------------------------------------------- #
# Import logic
# --------------------------------------------------------------------------- #
def import_workbook(path, verbose=True):
    wb = openpyxl.load_workbook(path, data_only=True)
    stats = {"sheets": 0, "rows": 0, "clients": 0, "projects_created": 0,
             "projects_skipped": 0, "requirements": 0}

    def get_client(name):
        key = extract_client(name) or "Unknown Client"
        c = Client.query.filter_by(name=key).first()
        if c is None:
            c = Client(name=key)  # type: ignore
            db.session.add(c)
            db.session.flush()
            stats["clients"] += 1
        return c

    for sheet_name, extractor in SHEET_EXTRACTORS:
        if sheet_name not in wb.sheetnames:
            if verbose:
                print(f"  ! sheet not found: {sheet_name}")
            continue
        ws = wb[sheet_name]
        stats["sheets"] += 1
        for row in extractor(ws):
            stats["rows"] += 1
            client = get_client(row["client"])
            job = clean(row["job"])
            name = clean(row["name"])
            method = SubmissionMethod.query.filter_by(
                code=row["method"]).first()
            if method is None:
                print(f"  ! missing submission method {row['method']}")
                continue

            # Dedupe: same job code (or same site under same client) once.
            if job:
                project = Project.query.filter_by(job_code=job).first()
            else:
                project = Project.query.filter_by(client_id=client.id,
                                                  name=name).first()
            if project is not None:
                stats["projects_skipped"] += 1
                continue

            project = Project(
                client_id=client.id, name=name, job_code=job or None,  # type: ignore
                submission_method_id=method.id,  # type: ignore
                common_criteria=clean(row["common"]) or None,  # type: ignore
                sds_criteria=clean(row["sds"]) or None,  # type: ignore
                target_date_rule=clean(row["target"]) or None,  # type: ignore
                step_narrative=row["narrative"] or None,  # type: ignore
                default_team_code=detect_team(row["narrative"]) or None,  # type: ignore
            )
            db.session.add(project)
            db.session.flush()
            stats["projects_created"] += 1

            for i, code in enumerate(detect_docs(row["docs_text"])):
                dt = DocumentType.query.filter_by(code=code).first()
                if dt is None:
                    print(f"  ! unknown doc type {code} for {name}")
                    continue
                db.session.add(ProjectDocumentRequirement(
                    project_id=project.id, doc_type_id=dt.id, sequence=i,  # type: ignore
                    required=True,  # type: ignore
                ))
                stats["requirements"] += 1

        db.session.commit()

    wb.close()
    return stats


def main():
    parser = argparse.ArgumentParser(description="Import Excel invoicing data")
    parser.add_argument("--reset", action="store_true",
                        help="Drop and recreate the database before importing")
    args = parser.parse_args()

    if args.reset:
        db_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "data", "sds.db")
        if os.path.exists(db_path):
            os.remove(db_path)
            print(f"Removed existing database: {db_path}")

    app = create_app()
    with app.app_context():
        print(f"Importing {EXCEL_FILE} ...")
        stats = import_workbook(EXCEL_FILE)
        print("\nImport complete:")
        for k, v in stats.items():
            print(f"  {k:18s}: {v}")

        # Quick sanity report
        print("\nSanity:")
        print(f"  Clients   : {Client.query.count()}")
        print(f"  Projects  : {Project.query.count()}")
        counts = {}
        for m in SubmissionMethod.query.all():
            counts[m.code] = Project.query.filter_by(
                submission_method_id=m.id).count()
        for code, n in counts.items():
            print(f"    {code:12s}: {n} projects")
        print(f"  Requirements: {ProjectDocumentRequirement.query.count()}")


if __name__ == "__main__":
    main()
